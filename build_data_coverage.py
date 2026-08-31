"""
Builds docs/RERA_Data_Coverage.xlsx -- what each authority and each workflow
actually yields, and where it lands in the Company Charter.

WHY THIS IS A SCRIPT AND NOT A SPREADSHEET
Kept as code so it is version-controlled, diffable in review, and cannot
drift silently: the FINDINGS below are the single source of truth and the
workbook is regenerated from them. Hand-editing the .xlsx is the one thing
NOT to do -- the next run overwrites it.

Every row records how the finding was established. "confirmed-live" means
this pipeline actually fetched it; "observed" means it was seen on the
portal but not yet fetched by code; "unaudited" means nobody has looked.
That distinction matters because ~20 of ~30 state portals are unaudited, and
a coverage matrix that hides its own uncertainty is worse than none.

    python build_data_coverage.py

Update it whenever a workflow learns something new -- add to FINDINGS, rerun.
"""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join("docs", "RERA_Data_Coverage.xlsx")

# --- house style ---------------------------------------------------------
_HEAD_FILL = PatternFill("solid", fgColor="1F3864")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")
_WRAP = Alignment(wrap_text=True, vertical="top")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Availability vocabulary, deliberately small. Colour carries the same
# meaning as the word so the sheet is readable at a glance and still
# correct when printed in greyscale.
_STATUS_FILL = {
    "Yes": PatternFill("solid", fgColor="C6EFCE"),
    "Partial": PatternFill("solid", fgColor="FFEB9C"),
    "No": PatternFill("solid", fgColor="FFC7CE"),
    "Not published": PatternFill("solid", fgColor="FFC7CE"),
    "Unaudited": PatternFill("solid", fgColor="D9D9D9"),
    "Not built": PatternFill("solid", fgColor="D9D9D9"),
}

_EVIDENCE_FILL = {
    "confirmed-live": PatternFill("solid", fgColor="C6EFCE"),
    "observed": PatternFill("solid", fgColor="FFEB9C"),
    "unaudited": PatternFill("solid", fgColor="D9D9D9"),
}


# =========================================================================
# WORKFLOW A -- RERA. What each authority publishes, per data item.
# =========================================================================
# Columns are authorities; the last two record where it lands in the Charter.
RERA_STATES = ["MahaRERA", "TG-RERA", "GujRERA", "K-RERA", "JHARERA", "WBRERA",
               "UP-RERA", "TNRERA", "HARERA", "Delhi-RERA"]

# How many of those the RERA_FINDINGS tuples below carry inline.
_INLINE_STATES = 6

# The four authorities added on 2026-08-24, keyed by data item rather than
# widened into the tuples below -- 21 carefully worded rows are not worth
# rewriting to add four columns, and a lookup makes an unfilled row loudly
# "Unaudited" instead of silently short.
#
# EVERY "Yes" HERE WAS SEEN ON A REAL RECORD, not inferred from a portal's
# menu: UPRERAPRJ14636 and UPRERAPRJ2499, TNRERA/29/BLG/0001/2026 and
# TN/16/Building/0001/2024, RERA-GRG-741-2020, and Delhi's whole register.
# Where nobody has looked the cell says "Unaudited", which is the honest
# answer and the one this sheet exists to keep visible.
RERA_FINDINGS_LATER = {
    # data item: (UP-RERA, TNRERA, HARERA, Delhi-RERA)
    "Project identity (name, status, type, dates)":
        ("Yes", "Yes", "Yes", "Partial"),
    "Registration number retrievable from the portal":
        ("Partial", "Yes", "Yes", "Yes"),
    "Promoter / partner identity":
        ("Yes", "Yes", "Yes", "Partial"),
    "Professionals of record (architect/engineer/CA)":
        ("Partial", "Partial", "Not published", "Not published"),
    "Land details / survey numbers":
        ("Yes", "Partial", "Partial", "Not published"),
    "Bank accounts (escrow / collection)":
        ("Yes", "Partial", "Yes", "Not published"),
    "Document library (downloadable files)":
        ("Yes", "Yes", "Yes", "Not published"),
    "Complaints register":
        ("Not published", "Yes", "Not published", "Yes"),
    "Appeals register":
        ("Partial", "Not published", "Yes", "Yes"),
    "Orders / judgments search":
        ("Not published", "Yes", "Not published", "Yes"),
    "Promoter's other projects (track record)":
        ("Partial", "Partial", "Yes", "Yes"),
    "Past-experience declarations":
        ("Not published", "Not published", "Not published", "Not published"),
    "AUDITED BALANCE SHEET":
        ("Not published", "Not published", "Yes", "Not published"),
    "Audited profit & loss statement":
        ("Not published", "Not published", "Not published", "Not published"),
    "Income-tax returns":
        ("Not published", "Not published", "Not published", "Not published"),
    "Defaulters list":
        ("Yes", "Yes", "Yes", "Partial"),
    "Projects under investigation":
        ("Partial", "Partial", "Not published", "Yes"),
    "Cost incurred vs estimated cost":
        ("Partial", "Partial", "Yes", "Not published"),
    "Delay reasons (promoter-declared)":
        ("Not published", "Not published", "Partial", "Not published"),
    "NOC expiry and renewal tracking":
        ("Not published", "Not published", "Not published", "Not published"),
    "Construction progress / QPR":
        ("Partial", "Yes", "Partial", "Partial"),
}

# Notes for the four new columns, appended to the row's existing note so the
# reasoning travels with the cell rather than living only in a commit.
RERA_NOTES_LATER = {
    "Professionals of record (architect/engineer/CA)":
        "HARERA is Not published, audited live 2026-08-31 across 7 real Gurugram-bench projects (detail ids "
        "1444, 1011, 2953, 1641, 1353, 1856, 2967) plus a full listing of every numbered section "
        "heading on the REP-I form: zero occurrences of 'architect' as a named professional on 6 of "
        "7, and the one project where 'architect' does appear is only the phrase 'AS PER ARCHITECT "
        "CERTIFICATE ATTACHED' inside a construction-status remark, naming no one. 'Chartered "
        "Accountant' appears on every project, but only as the signer of a non-default certificate "
        "and a balance-sheet compliance certificate -- a role, not a named professional-of-record "
        "entry with a licence number the way MahaRERA's professionals table carries one. HARERA's "
        "REP-I form simply asks for none of this: company/director details, land/FAR, approvals, "
        "construction and sales status, financials and bank accounts, statutory approvals -- no "
        "architect/engineer/CA identity section anywhere in the form's own structure.",
    "Past-experience declarations":
        "All three move from Unaudited to Not published, audited live 2026-08-31. UP-RERA: zero "
        "hits for 'experience'/'past project'/'completed project'/'track record' across four real "
        "detail pages (14636, 11528, 2499, plus two invalid ids returning the portal's own empty "
        "shell) -- matches the authority's own detail-page contents already documented in "
        "states/uttarpradesh.py (promoter identity, khasra grid, bank account, architect/engineer, "
        "documents, QPR certificates), which never mentions past experience either. TNRERA: zero "
        "hits across three real registrations spanning both the dynamic public-view2 project page "
        "and public-view1 promoter page (TN/29/Building/0328/2024) and a legacy static register "
        "entry (TN/16/Building/0001/2024) -- neither view's own section list (Financial Indicators, "
        "Promoter Detail, Approval Details, Project Bank Details, etc.) includes anything resembling "
        "past experience; this is a DIFFERENT data item from 'Promoter's other projects (track "
        "record)' above, which is about finding a promoter's OTHER TNRERA filings by name match, not "
        "a self-declared history filed with this application. HARERA: same REP-I heading sweep as "
        "Professionals of record above -- no section asks for it.",
    "Construction progress / QPR":
        "HARERA is Partial, audited live 2026-08-31 across 8 real Gurugram-bench projects: the REP-I "
        "form's own construction-status fields are real but inconsistent in shape by project type. "
        "Plotted colonies carry a numeric 'Percentage completion Upto the date of application' for "
        "infrastructure development (85% and 59.24% and 12.30% seen on three different projects) plus "
        "a per-block-type qualitative status ('CIVIL STRUCTURE COMPLETED', 'WE HAVE RECEIVED OC ON "
        "128 PLOTS'); group-housing/apartment projects instead carry only the qualitative narrative, "
        "no percentage. Both also carry an itemised estimated-vs-actual-expenditure table per "
        "infrastructure category (roads, water supply, drainage, electricity, etc.), which is a "
        "genuine progress signal in its own right. Stays Partial rather than Yes for two reasons: the "
        "field a reader gets depends on project type rather than being uniform, and there is no "
        "visible mechanism confirming these figures refresh on a recurring cadence rather than being "
        "frozen at the date of the last REP-I filing or amendment.",
    "Registration number retrievable from the portal":
        "UP-RERA is Partial: a legacy UPRERAPRJ number resolves with no search at all (the "
        "detail page id is the number's own numeric suffix), but the scheme used since ~2024 "
        "(UPRERAPRJ378870/03/2025) does not follow that and cannot be resolved here.",
    "Complaints register":
        "TNRERA publishes THREE order registers (Authority, single-member bench, Adjudicating "
        "Officer) and Delhi-RERA one, all naming complainant AND respondent in their own "
        "columns -- which is what makes them promoter-searchable. UP-RERA and HARERA both "
        "gate their case search behind a CAPTCHA keyed on a case number, so complaints there "
        "are UNKNOWN rather than absent.",
    "Promoter's other projects (track record)":
        "UP-RERA is now Partial rather than No: CONFIRMED LIVE 2026-08-26 via a human-solved "
        "CAPTCHA (up_captcha_search.py) that the search genuinely works, correctly returning "
        "UPRERAPRJ14636 (BALAJI GREENS) for BALAJIMAHIMA INFRATECH PRIVATE LIMITED in "
        "Barabanki district. It stays Partial, not Yes, because it demands a district before a "
        "promoter, so a full portfolio would need a solved CAPTCHA per district across 75 "
        "districts -- unattended automation is still refused (group_sweep._CANNOT_SEARCH), but "
        "a human-in-the-loop lookup is a real, working path now, not merely a theoretical one. "
        "TNRERA is Partial for a different reason: no promoter id and no promoter search, so a "
        "portfolio is a NAME match across year tables and every row is a candidate.",
    "Bank accounts (escrow / collection)":
        "UP-RERA publishes the project account number in full with bank and branch. TNRERA "
        "masks it (XXXXXXXXXX9995) while naming the bank, branch and the account holder.",
    "Land details / survey numbers":
        "UP-RERA publishes a KHASRA/PLOT grid with per-parcel areas and a registry/agreement "
        "grid. HARERA carries DTCP licence numbers rather than khasra numbers. Note these are "
        "the promoter's own declaration in every case, not an independent land record.",
    "Document library (downloadable files)":
        "UP-RERA's grid has a third state besides filed and missing: 7 of 31 rows on a real "
        "record carry the promoter's own 'NA', including the CA, ARCHITECT and ENGINEERS "
        "certificates -- not filed, which is a finding rather than a gap. HARERA links 60 "
        "documents on one Gurugram record.",
    "Project identity (name, status, type, dates)":
        "Delhi-RERA is Partial because there is no per-project record at all: the register's "
        "own View control is inert and every detail route probed returns nothing, so a Delhi "
        "project's identity is its register row and nothing more.",
    "Appeals register":
        "TNRERA is Not published, audited live 2026-08-26: a Tamil Nadu Real Estate Appellate Tribunal "
        "(TNREAT) DOES exist, distinct from the authority -- it has its own domain, "
        "tnreat.tn.gov.in, linked from TNRERA's own homepage -- but its public page carries no "
        "case search, cause list or judgment register at all, only a sign-in and a virtual-"
        "meeting-report link. Not a stub of the authority's own order registers: a genuinely "
        "separate body that simply publishes nothing searchable to the public. UP-RERA is "
        "Partial: its own Appellate Tribunal (UP-REAT) runs a live judgement search at "
        "efilingreat.up.gov.in/upreat/judgement.php with a free-text party-name field, "
        "CAPTCHA-gated. CONFIRMED LIVE 2026-08-26 the gate is passable, not just present: "
        "up_captcha_search.py opens a real browser, a human reads and solves the CAPTCHA, and "
        "the search genuinely runs -- tested against BALAJIMAHIMA INFRATECH PRIVATE LIMITED, "
        "which returned a clean 'No Data Found' in a real results table (Sr.No/Filing No./Case "
        "No./Case Title/Registration Date/Action), not an error or an empty page. Stays "
        "Partial rather than Yes because every search still costs one human CAPTCHA solve -- "
        "real and working, but not unattended. HARERA is Yes: the "
        "Haryana Real Estate Appellate Tribunal is served off HARERA's own domain "
        "(haryanarera.gov.in/admincontrol/judgements/3), a plain GET with no CAPTCHA -- 2,779 "
        "rows of Appeal No. / Appellant Name / Respondent Name / Date of Decision, confirmed "
        "promoter-searchable. Delhi-RERA is now Yes, upgraded from Partial 2026-08-26: REAT "
        "Delhi (shared with the UT of Chandigarh, not linked from any Delhi-RERA page, found "
        "only by web search) publishes a live 505-row register that names no party in its own "
        "columns -- but every row links a scanned judgement PDF, and REAT's own case captions "
        "name the Appellant and Respondent on the PDF's first page. "
        "states/adapter_delhi.py's build_appeal_party_index() downloads and OCRs (PyMuPDF "
        "native text first, Tesseract fallback -- confirmed live that every PDF sampled is a "
        "scan, zero native text) just that first page across all 481 distinct order PDFs behind "
        "the 505 rows, extracting both party names by the caption's own structure rather than "
        "guessing at fixed columns. Confirmed live 2026-08-26: 437/505 rows (87%) now carry a "
        "real party name -- searching 'Parsvnath' returns 27 real hits, e.g. 'Bimal Kumar & Ors. "
        "vs M/s Parsvnath Landmark Developers Pvt. Ltd.' Of the 68 rows that don't, 63 are the "
        "AUTHORITY'S OWN register linking a PDF that 404s (a genuine broken link on their side, "
        "confirmed directly), and only 5 PDFs (12 rows, all 2021-2022 vintage) use a caption "
        "shape the parser doesn't recognise. search_appeals_by_party() is the promoter-facing "
        "search this unlocks, still unwired into acquire().",
    "AUDITED BALANCE SHEET":
        "TNRERA is Not published, audited live 2026-08-26 against TNRERA/29/BLG/0001/2026, /0004/2026 and "
        "/0005/2026: the promoter view carries a 'Financial Indicators (Rs. in Lakhs)' block "
        "with a bare self-declared Net Worth NUMBER (populated on one of three: Rs "
        "2,00,00,000) but no uploaded balance-sheet document anywhere in the documents list on "
        "any of the three projects checked. UP-RERA is Not published: the document grid was read on three "
        "real projects (51 rows, 20 distinct labels) and none resembles a balance sheet -- the "
        "closest is the Form REG-3 CA certificate, which certifies compliance rather than filing "
        "the statement itself. HARERA's Yes is real but softer than GujRERA's or JHARERA's actual "
        "uploaded statements: it rests on a Part-D compliance declaration ('Annex copy of the "
        "balance sheet of last 3 years: Yes') and a CA-certificate document referencing 'BOOKS OF "
        "ACCOUNTS/ BALANCE SHEET' -- no document in the 60-row library checked is itself labelled "
        "'balance sheet'.",
    "Audited profit & loss statement":
        "TNRERA is Not published, same 'Financial Indicators' block and same three projects: 'Net Profit / "
        "Loss' is a bare self-declared number (Rs 72,16,670 on the one populated record), never "
        "a filed P&L statement. UP-RERA is Not published, same three-project document-grid read as above. "
        "HARERA is Not published: 'profit and loss'/'profit & loss' occurs zero times across four project "
        "detail pages checked (1444, 681, 3723, 431), against a single 'balance sheet' mention "
        "each -- the two are not filed together.",
    "Income-tax returns":
        "TNRERA is Not published, same block: 'Taxes Paid - IT (GST/ST)' is a single number that conflates "
        "income tax with GST/service tax (Rs 27,92,915 on the populated record) and no ITR "
        "document is filed anywhere in the three document sets checked. UP-RERA is Not published, same "
        "document-grid read. HARERA is Not published: 'income tax'/'ITR' occurs zero times across the same "
        "four pages.",
    "Defaulters list":
        "TNRERA is Yes, audited live 2026-08-26: https://rera.tn.gov.in/building/online/penalty "
        "and .../layout/online/penalty are live, unpaginated pages naming the promoter, address, "
        "project and penalty levied -- 1 row on Building (TNRERA/PBF/0092/2025, M/S. VIKAS "
        "MANTRA PROPERTIES & INFRASTRUCTURE PRIVATE LIMITED, Rs 20,10,940, 13-02-2025) and 146 "
        "on Layout. Titled 'Penalty', not 'Defaulters' or 'Black List' -- no register under "
        "either of those names was found, and no separate revoked/cancelled-registration list "
        "exists. An unwired parser (parse_penalty_register / fetch_penalty_notices) was added "
        "to states/adapter_tamilnadu.py, mirroring adapter_westbengal.fetch_defaulters(). UP-RERA "
        "is Yes: https://www.up-rera.in/DefaulterList (reachable only via the homepage's "
        "DE-REGISTERED PROJECTS postback, not a plain link) serves a live 72-row register -- reg. "
        "no., project name suffixed '(De-Registered Project)' or '(Defaulter Project)', district, "
        "promoter -- e.g. UPRERAPRJ7090, ANSAL PROPERTIES & INFRASTRUCTURE LIMITED. An unwired "
        "fetch_defaulters() was added to states/adapter_uttarpradesh.py. HARERA is Yes: "
        "/admincontrol/cancelled_projects/{bench} (menu label 'Defaulter/ Cancelled/ Suspended/ "
        "Abeyance Projects') is a distinct, live register -- 23 Panchkula + 5 Gurugram rows -- "
        "confirmed DIFFERENT from /admincontrol/lapsed_projects/{bench} (319 + 234 rows), which "
        "is only registrations whose validity date has passed, not a defaulter finding; both got "
        "unwired parsers in states/adapter_haryana.py. Delhi-RERA is Partial: the authority never "
        "uses the word 'defaulter', but courtview/ExecutionInOrderJudgementsAuthorityInfo -- "
        "7,493 live rows of orders referred for enforcement because the promoter ('Judgement "
        "Debtor') did not comply -- is the closest real equivalent; got an unwired parser in "
        "states/adapter_delhi.py.",
    "Projects under investigation":
        "TNRERA stays Partial, but 'static (non-searchable)' is fixed as of 2026-08-26: both "
        "PDFs are NATIVE TEXT (confirmed live, zero OCR needed, unlike Delhi-RERA's REAT scans) "
        "and adapter_tamilnadu.py's parse_enforcement_pdf()/search_enforcement_lists_by_name() "
        "reads them with pdfplumber's table extraction -- 35 rows off the 5-page SCN list, 1,502 "
        "off the 118-page caution list, confirmed live end-to-end (searching 'Sivakaminathan' "
        "returns the real row). What did NOT change, because it is a finding about the "
        "authority's publication rather than a scraping gap: neither list covers a REGISTERED "
        "project under suo-moto scrutiny, only the unregistered-building enforcement pipeline -- "
        "'Show Cause Notice issued for levy of penalty for non registration' and a 'Public "
        "Caution Notice' of buildings claimed as personal-use and therefore left unregistered. "
        "UP-RERA stays Partial for the opposite reason: the three adjacent registers (Abeyance, "
        "NCLT, Withdrawn Registration) named in the earlier audit are REAL pages, but their data "
        "loads via a WebService1.asmx POST that returned HTTP 500 on every parameter tried, "
        "including a real browser driving the page's own default (empty-parameter) call -- "
        "confirmed live 2026-08-26 this is the AUTHORITY'S OWN backend failing, not a request-"
        "shaping problem on this end, so nothing was built against a data source that cannot "
        "currently be read at all. HARERA is Not published: the only relevant menu item, 'Suo Motu "
        "(Projects) Cause List', is a hearing-schedule generator requiring a specific date and a "
        "CAPTCHA, not a browsable register; the only visible trace of suo-moto outcomes is "
        "inside HREAT's judgement text (case numbers containing 'MT' for Motu). Delhi-RERA is "
        "Yes: courtview/SuoMotoCases is a live, promoter-named, 1,797-row register of the "
        "authority's own-motion notices and orders -- the clearest 'projects under "
        "investigation' register found on any of the four new states; got an unwired parser in "
        "states/adapter_delhi.py.",
    "Delay reasons (promoter-declared)":
        "TNRERA is Not published, audited live against TN/29/Building/0328/2024 whose register row reads "
        "'Extension given upto 31.07.2026 Completed': its detail view (public-view2) has no "
        "delay-reason, extension-reason or revised-completion-date field across any of its 20 "
        "sections -- the extension text lives only as free text in the register's own Current "
        "Status column, never as a structured field, and never states WHY. UP-RERA is Not published: "
        "'delay'/'reason'/'extension'/'revised' were grepped across three project pages with "
        "zero hits for 'delay' or 'reason'; the only adjacent field is a bare revised valid-upto "
        "date with no promoter-declared justification. HARERA is Partial: the project detail "
        "page carries structured 'Initial date of completion' and 'Likely date of completion' "
        "fields -- a computable delay signal -- but 'reason'/'justification' occurs zero times "
        "anywhere on the page, so there is no promoter-declared free-text reason, only the two "
        "dates.",
    "NOC expiry and renewal tracking":
        "TNRERA is revised from Partial to Not published, audited live 2026-08-26: the detail view does "
        "carry a 'Clearance / NOC Details' section, but its only fields are 'Clearance Type' "
        "(e.g. 'PWD & RS') and 'Uploaded Document' -- confirmed on both a fresh 2026 "
        "registration and an already-extended 2024 one -- with no expiry date or renewal-status "
        "field anywhere in either. UP-RERA is Not published: zero genuine 'NOC' occurrences across three "
        "project pages and up-rera.in's full 153-link menu (the only raw-HTML hits were inside "
        "the base64 __VIEWSTATE blob). HARERA is Not published: the 'Statutory Approvals Status' table "
        "lists licence/clearance/NOC numbers each marked 'ALREADY BEEN OBTAINED' with a date -- "
        "but that date is when it was obtained, not an expiry or validity-end date, and no "
        "separate NOC-validity register exists.",
}

RERA_FINDINGS = [
    # (data item, MH, TG, GJ, KA, JH, WB, charter facts field, note)
    ("Project identity (name, status, type, dates)",
     "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "rera_core_fields", ""),
    ("Registration number retrievable from the portal",
     "Yes", "Not published", "Yes", "Yes", "Yes", "Yes", "rera_core_fields.registration_number",
     "TG-RERA's public record does not display its own registration number; search is by project name."),
    ("Promoter / partner identity",
     "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "corporate_identity",
     "K-RERA publishes partner PANs and land-owner SHARES -- neither appears on a MahaRERA record."),
    ("Professionals of record (architect/engineer/CA)",
     "Yes", "Not published", "Yes", "Yes", "Yes", "Yes", "local_planning.professionals_of_record",
     "GujRERA splits these across englist/calist/acrchlist/contr; the adapter normalises to one list. "
     "JHARERA is the only authority here that files a professional's PAN as a plain text field -- for "
     "the contractor, architect and structural engineer -- rather than a scanned card."),
    ("Land details / survey numbers",
     "Yes", "Yes", "Yes", "Yes", "Yes", "Partial", "land_identification",
     "K-RERA gives per-owner shares against survey numbers. WBRERA is now Partial, built and "
     "confirmed live 2026-08-31: Land Area and a location block (address/district/block/police "
     "station/pincode) are structured and present on every project sampled, via "
     "adapter_westbengal.land_details(). A dag/mouza/J.L. survey reference is NOT structured -- it "
     "is free text sometimes embedded in the project address, present on roughly a third of a "
     "random six-district sample and absent from the rest, so land_reference_present says which "
     "case a given project is rather than reading empty."),
    ("Bank accounts (escrow / collection)",
     "Yes", "Yes", "Yes", "Yes", "Yes", "Not published", "rera_compliance.collection_account",
     "TG-RERA publishes the 100% collection account but leaves the 70/30 split accounts blank. "
     "WBRERA is revised from Not built to Not published, confirmed live 2026-08-31 across a random "
     "sample of six projects spanning six districts and registration years: 'account', 'escrow' and "
     "'IFSC' occur zero times on any of them (the only 'bank' hits are the district-menu entry for "
     "Bankura). This is the authority never publishing the field at all, not an adapter gap -- "
     "adapter_westbengal.py now states this outright in _AUTHORITY_NOTES rather than leaving it to "
     "look like unfinished work."),
    ("Document library (downloadable files)",
     "Yes", "Not published", "Yes", "Yes", "Yes", "Yes", "document_library",
     "GujRERA 42/42 retrieved. K-RERA 112/152 -- the other 40 are LISTED but the portal serves 0 bytes. "
     "JHARERA labels every document link 'View'; document_label() derives the real name from the "
     "surrounding table or header. WBRERA's links already carry real text, so no derivation is needed."),
    ("Complaints register",
     "Yes", "Not published", "Not published", "Yes", "Partial", "Not published",
     "rera_core_fields.total_complaints_count",
     "K-RERA: use the STATE-WIDE /projectComplaintReport. The per-project page is NOT reliable -- it "
     "showed no complaints for a project the register lists with 12. JHARERA has no per-project count "
     "either; the adapter name-matches the promoter against the state-wide disposed-complaint "
     "register, which yields POSSIBLE matches to confirm, not a confirmed count. WBRERA publishes no "
     "complaint register through this interface at all -- recorded as None, never 0."),
    ("Appeals register",
     "Yes", "Not published", "Not published", "Not published", "Not published", "Not published",
     "rera_core_fields.total_appeals_count", ""),
    ("Orders / judgments search",
     "Yes", "Not published", "Not published", "Yes", "Yes", "Partial", "sources[] (topic=litigation)",
     "K-RERA now reads FIVE promoter-keyed registers (order search, authority orders, AO orders, "
     "interim orders, complaints under process) -- 15,600+ rows including a penalty table naming the "
     "violation, section and amount. JHARERA's judgement/order register is searchable by promoter, "
     "though a share of rows name both parties in one unsplit column and are therefore invisible to "
     "that search. WBRERA's order register names no party at all; the join runs through its cause "
     "lists instead, so its coverage note matters more than its row count."),
    ("Promoter's other projects (track record)",
     "Yes", "Not published", "Not published", "Yes", "Yes", "Not published", "promoter_portfolio",
     "K-RERA embeds the WHOLE state index client-side (9,888 projects with promoter names), so a "
     "portfolio is one request. GujRERA confirmed absent. WBRERA publishes no promoter search and its "
     "state index does not name the promoter either, so a portfolio there would mean opening all "
     "~4,700 project pages -- the adapter returns None and says why rather than sampling a partial set."),
    ("Past-experience declarations",
     "Yes", "Not published", "Yes", "Not published", "Yes", "Yes", "promoter_portfolio.totals", ""),
    ("AUDITED BALANCE SHEET",
     "Not published", "Not published", "Yes", "Partial", "Yes", "Yes", "(unmapped -- see Charter Mapping sheet)",
     "GujRERA findoc block. MahaRERA publishes nothing equivalent. Highest-value differential finding. "
     "JHARERA also files an audited balance sheet as a downloadable document, labelled directly on "
     "the project page. WBRERA is now Yes, confirmed live 2026-08-31 across 12 real projects checked: "
     "the application's own item 2 ('Audited balance sheet of the promoter for the preceding financial "
     "year and income tax returns... for three preceding financial years') carries real, downloaded "
     "files on every one -- e.g. 'Audited_Balance Sheet.pdf', 'Audited Balance Sheet 2024.PDF' -- "
     "already fetched by adapter_westbengal.py's existing document download, just never previously "
     "flagged as filed. Labelling is inconsistent enough (misspellings like 'blance sheet', combined "
     "balance-sheet+ITR files) that a downstream reader should still open the document rather than "
     "trust the label alone."),
    ("Audited profit & loss statement",
     "Not published", "Not published", "Yes", "Not published", "Not published", "Not published", "(unmapped)",
     "GujRERA findoc block. WBRERA's own disclosure item asks for a balance sheet and income-tax "
     "returns ONLY, never a P&L statement by name -- checked live across the same 12 projects, no "
     "filename or label suggests one is filed separately, so this stays Not published rather than "
     "assuming a P&L rides along inside one of the combined financial-statement PDFs."),
    ("Income-tax returns",
     "Not published", "Not published", "Yes", "Not published", "Yes", "Yes", "(unmapped)",
     "GujRERA findoc block. JHARERA files three years of income-tax returns per project, also as "
     "downloadable documents."),
    ("Defaulters list",
     "Not published", "Not published", "Not published", "Partial", "Partial", "Partial", "(unmapped)",
     "K-RERA /viewDefaultProjects -- observed, not yet built. JHARERA's REJECTED and SURRENDERED "
     "registers are the closest equivalent: their URLs are imported into the adapter but nothing "
     "fetches them yet. WBRERA's defaulters list (17 rejected/defaulting applications by name) IS "
     "parsed by adapter_westbengal.fetch_defaulters(), but that function is called from nowhere in "
     "acquire() or the litigation sweep -- written, untested against a live page, and unused."),
    ("Projects under investigation",
     "Not published", "Not published", "Not published", "Partial", "Not published", "Not published", "(unmapped)",
     "K-RERA /unregProjectList -- observed, not yet built."),
    ("Cost incurred vs estimated cost",
     "Not published", "Not published", "Partial", "Yes", "Not published", "Not published", "(unmapped)",
     "K-RERA publishes both, per particular. Direct input to the financial-strength sub-metric."),
    ("Delay reasons (promoter-declared)",
     "Not published", "Not published", "Not published", "Yes", "Not published", "Not published", "(unmapped)",
     "K-RERA detail page carries a delay-reason table."),
    ("NOC expiry and renewal tracking",
     "Not published", "Not published", "Not published", "Yes", "Not published", "Not published", "(unmapped)",
     "K-RERA tracks NOC validity dates and whether renewed."),
    ("Construction progress / QPR",
     "Yes", "Partial", "Yes", "Yes", "Yes", "Not published", "rera_compliance.construction_progress",
     "JHARERA is now Yes, audited live 2026-08-31 against a real project (jharera.jharkhand.gov.in "
     "id 2625, PRANAMI BUILDERS): the flat/floor grid the adapter already reads for sold/unsold "
     "counts (adapter_jharkhand.py's own 'units' table, matched on 'flat no'+'sold status') carries "
     "its OWN per-unit 'Completion Status' column alongside Sold Status -- 101 rows, one per flat, "
     "each independently marked (every one 'Under Process' on this project, since it is still under "
     "construction). Real, structured, granular data that was already being fetched into `units` but "
     "never surfaced as a construction-progress finding. WBRERA is now Not published, audited live "
     "2026-08-31: the project detail page (Crown, WBRERA/P/NOR/2025/002592) carries only SITE-WIDE "
     "marquee notices and a QPR_User_Manual.pdf link telling promoters a Quarterly Project Status "
     "Update is mandatory -- the mechanism exists as a filing requirement, but no per-project "
     "percentage, status or progress field is ever rendered on the page itself or in its document "
     "library. THE JHARERA PORTAL ITSELF IS UNSTABLE, not merely slow -- worth flagging beyond this "
     "one row. During the original audit 3 of 4 fetch attempts against the same URL (id 2625) timed "
     "out before one succeeded and produced the finding above. A follow-up retry session immediately "
     "after went 0 for 9 across three separate batches (40-50s timeouts, including three straight "
     "attempts at the exact same id that had worked minutes earlier) -- the portal was not reachable "
     "at all during that window. So JHARERA's Yes here rests on ONE successful live fetch, not a "
     "repeatable check; a future run finding this authority's portal unresponsive is the portal's own "
     "instability, not evidence the field is absent, and should not be read as a regression."),
]

# =========================================================================
# WORKFLOW B -- CTS / land records
# =========================================================================
CTS_FINDINGS = [
    # (field, source, availability, cost, evidence, note)
    ("Owner / holder name", "Maha Bhulekh Property Card (हक्काचा मूळ धारक)",
     "Partial", "Free", "confirmed-live",
     "PRESENT on the free card. Extraction is what is broken -- the parser looks for ZaubaCorp's li.row pattern, but the card is an HTML table."),
    ("Area", "Maha Bhulekh Property Card (क्षेत्र चौ.मी.)",
     "Partial", "Free", "confirmed-live", "Present on the card; same extraction gap."),
    ("Tenure / occupancy class", "Maha Bhulekh Property Card (धारणाधिकार)",
     "Partial", "Free", "confirmed-live", "Only source for this field -- Index II does not carry occupancy class."),
    ("Encumbrance (on the card)", "Maha Bhulekh Property Card (इतर भार)",
     "Partial", "Free", "confirmed-live", "Present as a labelled row."),
    ("Mutation entries", "Maha Bhulekh Property Card (फेरफार table)",
     "Partial", "Free", "confirmed-live", "Full dated table with mutation numbers, present on the card."),
    ("Registered deeds / parties / consideration", "IGR Maharashtra e-Search (freesearchigrservice.maharashtra.gov.in)",
     "Yes", "Free + CAPTCHA", "confirmed-live",
     "Audited live 2026-09-01 with a REAL result, not just a reachable form. Two search modes, both "
     "free and CAPTCHA-gated (confirmed via DOM inspection, not just a page-text guess): 'Property "
     "Details' (Year -> District -> Village/Area, an autocomplete text field paired with a dropdown "
     "-> Survey/CTS/Milkat/Gat/Plot No.; a name-based search is also offered once a property number "
     "is entered) and 'Document Number' (Registration Type [eFiling/eRegistration/Regular/iSarita "
     "2.0] -> District, all 37 -- this mode is genuinely statewide, not Mumbai-only -> SRO -> Year -> "
     "Doc No.). Document Number/SRO Mumbai 9 (Andheri)/2024/#100 returned a real registered Leave & "
     "License agreement dated 03/01/2024 -- Seller/Lessor 'Ramesh Babulal Shah', Purchaser/Lessee "
     "'M/s Matushri Impex' (through partner Ramesh T. Bhalani), full property description (Bandra "
     "Kurla Complex premises, CTS No. 4207) AND consideration (monthly rent Rs 44,100 plus Rs "
     "5,29,200 advance for the first 12 months) -- all present in the RESULT LIST ROW ITSELF, no "
     "further click needed. Each row also carries an 'IndexII' button (a further postback, "
     "__doPostBack('RegistrationGrid','indexII$0')) documented in the site's own User Guide as "
     "opening a fuller Index II view in a separate window -- present and clickable, but its own "
     "content was NOT captured this pass (the popup did not materialize in this environment), so "
     "that specific view stays unconfirmed rather than assumed. A real, stated limitation from the "
     "site's own FAQ: Power of Attorney and Will deeds are NOT available in search results. Mumbai "
     "city and Suburban districts go back to 1985; 'selected offices' elsewhere now reach back before "
     "the earlier 2002 baseline too, per the site's own notice."),
    ("Equitable mortgages", "IGR Notice of Intimation (freesearchigrservice.maharashtra.gov.in)",
     "Yes", "Free + CAPTCHA", "confirmed-live",
     "The MECHANISM is confirmed live, sharing the exact same interface, schema and evidence as the "
     "row above -- 'Registration Type' (eFiling/eRegistration/Regular/iSarita 2.0) is a filing-"
     "channel category, not a deed-type filter, so a Notice of Intimation is just one of many "
     "possible 'DName' (deed name) values a search can return, discoverable the same way any other "
     "deed is: by its own document number, or by opening a property's full document list. No live "
     "example of an actual Notice of Intimation was captured this pass -- the one real result pulled "
     "was a Leave & License agreement -- so this confirms the CHANNEL, not a specific mortgage "
     "instance. Separately confirmed: the 'CERSAI Search' button on this same page is an INTEGRATED "
     "postback on igr's own portal (__doPostBack with a distinct 'aa' validation group), not an "
     "external redirect to cersai.org.in -- worth knowing before assuming a CERSAI cross-check means "
     "leaving this site. 'Compulsory in MAHARASHTRA AND GUJARAT ONLY, which is why CERSAI is not "
     "needed for those two states' is the pre-existing legal claim this row already carried; it was "
     "not independently re-verified this pass, since the portal itself doesn't state its own legal "
     "basis anywhere in the UI."),
    ("Lender, amount, satisfaction status", "MCA charge filings via ZaubaCorp",
     "Yes", "Free -- page already fetched", "confirmed-live",
     "STALE ROW, RECONCILED 2026-09-01: this said 'observed'/'NOT currently parsed' months after it "
     "actually was -- company_profile_check.charges was wired up and confirmed live on 2026-08-19 "
     "(commit d0536e6), and Sheet C's own 'MCA CHARGE FILINGS' row has said confirmed-live since then. "
     "This row just never got updated to match. See that row for the live evidence (4 open charges, "
     "~Rs 90.3 crore, HDFC Bank and Catalyst Trusteeship) -- kept here too, duplicated rather than "
     "deleted, because this is the only independent check on a promoter's declared mortgage and a "
     "reader scanning the land-records sheet for that fact should not be told it is still unbuilt."),
    ("Karnataka land records (Bhoomi)", "landrecords.karnataka.gov.in, RTC and Mutation Copy view",
     "Yes", "Free + CAPTCHA", "confirmed-live",
     "Audited live 2026-09-01. District -> Taluk -> Hobli -> Village drill-down, then survey number, "
     "then 'Fetch Survey Details' -- no login. A CAPTCHA text field (id=captchatextbox) IS present, "
     "confirmed by inspecting the live DOM directly rather than trusting a text summary of the page: "
     "an earlier automated read of this same portal claimed 'no explicit CAPTCHA mentioned', which was "
     "wrong -- the CAPTCHA is rendered by client-side JS and invisible to a fetch that doesn't run it."),
    ("Gujarat land records (AnyROR)", "anyror.gujarat.gov.in, Rural Land Record",
     "Yes", "Free + CAPTCHA", "confirmed-live",
     "Audited live 2026-09-01. One page, one District -> Taluka -> Village form, but a 15-option record-"
     "type dropdown behind it: VF-7 (survey details), VF-8A (khata/account), VF-6 (hakka patrak -- "
     "rights/mutation entries), 135-D mutation notice, revenue case details, and 'KNOW KHATA BY OWNER "
     "NAME' -- an owner-NAME search, confirmed present as a dropdown option though not opened this pass. "
     "A CAPTCHA text field (id=ContentPlaceHolder1_txt_captcha_1) is present, same DOM-inspection "
     "correction as Karnataka above -- a page-text summary alone also missed this one."),
    ("Uttar Pradesh land records (Bhulekh + eKhasra)", "upbhulekh.gov.in (khatauni/ROR) and "
     "ekhasra.up.gov.in (khasra/crop survey) -- two separate systems, not one",
     "Partial", "Free + CAPTCHA (khatauni) / Free (eKhasra geography step)", "confirmed-live",
     "Audited live 2026-09-01. upbhulekh.gov.in's khatauni (Record of Rights) copy view is behind a "
     "CAPTCHA text field (name=captcha) on the homepage's own District/Tehsil/Pargana form; the site "
     "ALSO has a separate full account login ('खतौनी लॉगिन' / 'रियल टाइम खतौनी लॉगिन') for a fuller "
     "record, not audited here. ekhasra.up.gov.in is a DIFFERENT system for crop/khasra survey data "
     "(fasli-year crop entries, tree counts), not land ownership -- its District -> Tehsil -> Village "
     "selection step showed no CAPTCHA element in the DOM, but the final record-display step past "
     "village selection was not reached this pass, so that step is unconfirmed rather than assumed "
     "clean. Partial because the two systems only jointly cover part of what a 7/12-equivalent needs.",
     ),
    ("Telangana land records (Bhu Bharati, successor to Dharani)",
     "bhubharati.telangana.gov.in -- dharani.telangana.gov.in (the name this codebase has referenced) "
     "is retired",
     "No", "Requires a mobile-OTP citizen account, not free-and-anonymous", "confirmed-live",
     "Audited live 2026-09-01. dharani.telangana.gov.in itself refused every connection tried (browser "
     "navigation AND a direct fetch both failed) -- not a network fluke on this end: Telangana's Revenue "
     "Department retired Dharani and replaced it with Bhu Bharati under the Telangana Bhu Bharati "
     "(Record of Rights in Land) Act, 2025, confirmed via web search, so this codebase's own comments "
     "citing 'Dharani (TG)' are now stale. bhubharati.telangana.gov.in DOES load, but every visible "
     "'Transactional Services' and 'Information Services' tile on its homepage routes to /Citizen, a "
     "login gate asking for Mobile No. + Password/OTP with self-registration ('New user please Sign "
     "Up') -- no anonymous district/survey-number lookup path was found anywhere on the public site, "
     "unlike Karnataka, Gujarat or Uttar Pradesh above. A promoter's Telangana land details are "
     "therefore not reachable by this pipeline without an Indian mobile number to receive an OTP, which "
     "is a materially different access model from the other three states' CAPTCHA-only public search."),
]

# =========================================================================
# WORKFLOW C -- CIN / corporate identity. Already national.
# =========================================================================
CIN_FINDINGS = [
    ("Legal name, CIN/LLPIN, status, class, category, ROC", "ZaubaCorp / Tofler / InstaFinancials",
     "Yes", "Free", "confirmed-live", "company_profile_check",
     "Merged across three mirrors, first non-empty wins. ROC is read from the record, never assumed."),
    ("Incorporation date, registered address", "MCA mirrors", "Yes", "Free", "confirmed-live",
     "company_profile_check", ""),
    ("Authorised and paid-up capital", "MCA mirrors", "Yes", "Free", "confirmed-live",
     "company_profile_check", "Paywall-guarded: _looks_paywalled discards upsell copy rather than rendering it."),
    ("Current and past directors (DIN, designation, dates)", "ZaubaCorp", "Yes", "Free", "confirmed-live",
     "company_profile_check.current_directors", "Merged by DIN; disagreements surface as roster_conflicts."),
    ("Insolvency status", "IBBI (official)", "Yes", "Free", "confirmed-live",
     "ibbi_insolvency_check",
     "Caveat recorded in code: a fake identifier returns the same 'clean' page, so absence is not proof of existence."),
    ("Credit rating", "ICRA + Infomerics", "Yes", "Free", "confirmed-live",
     "credit_rating_check", "CRISIL/CARE not implemented. Exact legal-name match only, never fuzzy."),
    ("Group / affiliated companies", "ZaubaCorp director & address crosswalk", "Yes", "Free", "confirmed-live",
     "group_companies_check", "Hard links only: shared director, shared registered office, or filed subsidiary/associate/JV."),
    ("MCA CHARGE FILINGS (lender, amount, assets, satisfaction)", "ZaubaCorp (same page)",
     "Yes", "Free", "confirmed-live", "company_profile_check.charges",
     "NOW PARSED. Live on a real promoter: 4 OPEN charges, ~Rs 90.3 crore, to HDFC Bank and "
     "Catalyst Trusteeship. No closure date = live encumbrance. The only independent check on "
     "a promoter's declared mortgage -- the RERA record gives an area and never a lender."),
    ("Balance sheet / P&L", "MCA (paid) / Tofler / InstaFinancials (paywalled)",
     "No", "Paid", "confirmed-live", "n/a",
     "NOT freely available for a private SPV -- EXCEPT in Gujarat, where GujRERA publishes audited statements per project."),
    ("Credit-rating rationale (revenue, debt, net worth)", "CARE / India Ratings / Infomerics",
     "Yes", "Free", "confirmed-live", "credit_rating_check.promoter.ratings[].rationale_url",
     "WIRED IN 2026-09-01, not just found: CARE, India Ratings and Infomerics all embed a path to their "
     "own rationale document in a response this pipeline already fetches for the bare rating -- "
     "Infomerics nests a PDF url four levels into a current instrument's own JSON; CARE's search "
     "response carries a CommonContent list of PDF filenames (date-sorted and Title-filtered here, "
     "since it mixes in group affiliates' filings under one shared CompanyID, unsorted); India Ratings "
     "renders its rationale as rich HTML, not a PDF, on its own page. All three confirmed by actually "
     "downloading/opening a real document -- Infomerics' carried a full 'Financials (Standalone)' "
     "table (Total Debt, Tangible Net Worth, EBITDA, PAT across two fiscal years); CARE's carried "
     "real Total Income/Net Worth/Gearing figures; India Ratings' page went further still -- bookings, "
     "collections, cash-flow-from-operations, forecast leverage ratios, a Strengths/Weaknesses key-"
     "rating-drivers section. company_charter.py's rating-comparison table now cites the rationale "
     "URL alongside the rating itself, and it is registered as its own sources[] entry. ICRA is the "
     "one gap left: a real per-entity rationale mechanism exists (/Rationale/ShowRationaleReport?Id=) "
     "but resolving which numeric Id belongs to a given company needs more reverse-engineering than "
     "this pass did. CRISIL is unchanged -- its own factsheet page already carried a short inline "
     "excerpt (the `rationale` key), not a document link, before this pass."),
]

# =========================================================================
# Charter mapping -- what is available but NOT yet used
# =========================================================================
UNMAPPED = [
    ("Audited balance sheet / P&L / ITR", "GujRERA findoc", "Gujarat projects only",
     "Financial strength sub-metric currently scores None for every project. This would score it for GJ.",
     "High"),
    ("Group entity graph (confirmed vs proposed)", "group_entities.py", "All corporate promoters",
     "BUILT. Propose by brand name, confirm by shared director / office / filed relationship. "
     "Link strength tiered -- address-only links excluded from sweeps.", "High"),
    ("K-RERA defaulters list", "K-RERA, observed", "Karnataka projects",
     "No equivalent exists in any other authority. Direct red-flag input.",
     "High"),
    ("K-RERA projects under investigation", "K-RERA, observed", "Karnataka projects",
     "Same -- a regulator-declared adverse signal.", "Medium"),
    ("K-RERA cost incurred vs estimated", "K-RERA detail page", "Karnataka projects",
     "Financial-strength scores None everywhere today. This is a direct, free input.", "High"),
    ("K-RERA delay reasons + NOC expiry", "K-RERA detail page", "Karnataka projects",
     "Promoter-declared delay reasons and lapsed-NOC tracking; no equivalent elsewhere.", "Medium"),
    ("Maha Bhulekh card fields", "Free portal, extraction broken", "Maharashtra projects",
     "Owner, area, tenure, encumbrance and mutation are all ON the card and none reach the document.",
     "High"),
    ("IGR Maharashtra registered-deed search", "igr_maharashtra_search.py -- BUILT 2026-09-01, "
     "human-in-the-loop like up_captcha_search.py/cts_resolve.py, not wired into run_company_charter()",
     "Maharashtra projects",
     "An independent, free, party-named check on registrations against a property/CTS number or a "
     "specific document number -- CAPTCHA-gated but otherwise unauthenticated. search_by_document_"
     "number() is confirmed against a real pulled result: seller/purchaser names, full property "
     "description and the ACTUAL CONSIDERATION AMOUNT in one row, no further click needed -- a direct "
     "corroboration (or contradiction) of a promoter's declared land dealings that nothing else in "
     "this pipeline checks independently. search_by_property() is also built (form mechanics "
     "confirmed live), but its OWN result table shape is not yet confirmed against a real row, unlike "
     "document-number search. Also carries an integrated CERSAI cross-search on the same page, not "
     "yet driven by either function. A standalone script, run by a human, same as every other "
     "CAPTCHA-gated tool here -- not called from the automated Charter pipeline.",
     "High"),
    ("JHARERA audited balance sheet + 3 years ITR", "adapter_jharkhand.py document library (labelled, downloaded)",
     "Jharkhand projects",
     "Same gap as GujRERA's findoc block: the financial-strength sub-metric scores None for every "
     "project regardless of state. This would score it for JH too.", "High"),
    ("JHARERA rejected + surrendered registration lists",
     "REJECTED_LIST / SURRENDERED_LIST, imported into the adapter but never fetched", "Jharkhand projects",
     "A rejected or surrendered registration is diligence material with no equivalent on MahaRERA. "
     "The URLs are already wired into the import list; nothing calls them yet.", "Medium"),
    ("WBRERA defaulters list", "adapter_westbengal.fetch_defaulters() -- written, unwired, untested live",
     "West Bengal projects",
     "17 rejected/defaulting applications keyed by name -- a cheap, direct red-flag input. The parser "
     "exists but is called from nowhere in acquire() or the litigation sweep.", "Medium"),
    ("UP-RERA de-registered/defaulter list", "adapter_uttarpradesh.fetch_defaulters() -- written, unwired, "
     "live-verified 2026-08-26 (72 rows)", "Uttar Pradesh projects",
     "Named by promoter outright -- ANSAL PROPERTIES & INFRASTRUCTURE LIMITED is on it. Reachable only "
     "via an ASP.NET postback, not a plain URL, which is why nothing had fetched it before.", "High"),
    ("HARERA cancelled/defaulter projects", "adapter_haryana.fetch_defaulter_projects() -- written, unwired, "
     "live-verified 2026-08-26 (23 Panchkula + 5 Gurugram)", "Haryana projects",
     "Distinct from the already-imported LAPSED_PROJECTS URL, which is validity expiry, not an "
     "authority action -- conflating the two would understate lapsed rows as defaults or vice versa.",
     "Medium"),
    ("TNRERA penalty register", "adapter_tamilnadu.fetch_penalty_notices() -- written, unwired, "
     "live-verified 2026-08-26 (147 rows across Building + Layout)", "Tamil Nadu projects",
     "Names promoter, project and the penalty amount levied -- TNRERA's own closest equivalent to a "
     "defaulters list, titled 'Penalty' rather than that.", "Medium"),
    ("TNRERA unregistered-project enforcement PDFs", "adapter_tamilnadu.search_enforcement_lists_by_name() "
     "-- written, unwired, live-verified 2026-08-26 (35 + 1,502 rows, native-text PDFs, no OCR needed)",
     "Tamil Nadu projects/promoters",
     "The show-cause and personal-use-caution lists, now actually searchable by name rather than merely "
     "linked as static PDFs. Ceiling is real, not fixable in code: both cover UNREGISTERED sites, never a "
     "registered project under suo-moto scrutiny.", "Medium"),
    ("Delhi-RERA suo-moto register", "adapter_delhi.parse_suo_moto_register() -- written, unwired, "
     "live-verified 2026-08-26 (1,797 rows)", "Delhi projects",
     "Names a Respondent/Promoter and project per row -- the clearest 'projects under investigation' "
     "signal found on any of the four newest states, and Delhi otherwise has no per-project record at "
     "all.", "High"),
    ("Delhi-RERA execution register", "adapter_delhi.parse_execution_register() -- written, unwired, "
     "live-verified 2026-08-26 (7,493 rows)", "Delhi projects",
     "Orders referred for enforcement because the promoter (named as Judgement Debtor) did not comply "
     "-- the authority's own closest equivalent to a defaulters list.", "Medium"),
    ("Delhi-RERA Appellate Tribunal (REAT) register, WITH party names",
     "adapter_delhi.build_appeal_party_index() / search_appeals_by_party() -- written, unwired, "
     "live-verified 2026-08-26 (437/505 rows named, 481 PDFs OCR'd)", "Delhi projects",
     "The register itself names no party -- not linked from any Delhi-RERA page either, found "
     "only by web search -- but every row links a scanned judgement PDF whose own case caption "
     "does. OCRing just the PDFs' first pages (PyMuPDF native text, confirmed zero on every "
     "sample -- Tesseract fallback) turned an unsearchable register into one where "
     "search_appeals_by_party('Parsvnath') returns 27 real hits. The only gaps are the "
     "authority's own 63 broken PDF links and 5 older (2021-2022) PDFs in a caption shape not "
     "yet recognised.", "High"),
]


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = _BORDER


def _finish(ws, widths, header_row, first_data_row, ncols):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row, max_col=ncols):
        for cell in row:
            cell.alignment = _WRAP
            cell.border = _BORDER
    ws.freeze_panes = ws.cell(row=first_data_row, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{ws.max_row}"


def _title(ws, text, subtitle, ncols):
    ws["A1"] = text
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, size=9, color="595959")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


def build_rera_sheet(wb):
    ws = wb.create_sheet("A - RERA by authority")
    ncols = 3 + len(RERA_STATES)
    _title(ws, "Workflow A: RERA -- what each authority publishes",
           "Yes / Partial / Not published / Unaudited. 'Not published' replaced the earlier separate "
           "'No' label 2026-08-31 -- both meant the same thing here (this authority does not publish "
           "this data item) and were used inconsistently row to row. 'Unaudited' means nobody has "
           "looked yet -- ~20 of ~30 state portals are in that state, and so are whole data items on "
           "the four newest authorities.", ncols)
    header = ["Data item"] + RERA_STATES + ["Charter facts field", "Note"]
    ws.append([])
    ws.append(header)
    header_row = ws.max_row
    _style_header(ws, header_row, ncols)
    unknown = set(RERA_FINDINGS_LATER) - {row[0] for row in RERA_FINDINGS}
    assert not unknown, f"RERA_FINDINGS_LATER names rows that do not exist: {sorted(unknown)}"
    for item, *rest in RERA_FINDINGS:
        # A row with no entry in the later table is Unaudited for those four,
        # never silently short -- the whole point of the sheet.
        later = RERA_FINDINGS_LATER.get(item, ("Unaudited",) * 4)
        states_vals = list(rest[:_INLINE_STATES]) + list(later)
        field, note = rest[_INLINE_STATES], rest[_INLINE_STATES + 1]
        extra = RERA_NOTES_LATER.get(item)
        if extra:
            note = f"{note} {extra}".strip()
        ws.append([item] + list(states_vals) + [field, note])
        for offset in range(len(RERA_STATES)):
            cell = ws.cell(row=ws.max_row, column=2 + offset)
            cell.fill = _STATUS_FILL.get(cell.value, _STATUS_FILL["Unaudited"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
    _finish(ws, [46] + [13] * len(RERA_STATES) + [30, 62], header_row, header_row + 1, ncols)


def build_cts_sheet(wb):
    ws = wb.create_sheet("B - CTS land records")
    ncols = 6
    _title(ws, "Workflow B: CTS / land records",
           "Maharashtra: the free portal already returns every field, and the extraction is what is "
           "broken -- no paid source is needed. Karnataka, Gujarat and Uttar Pradesh's own land-record "
           "portals are also free, CAPTCHA-gated, none built yet. Telangana is the exception: its own "
           "portal requires a mobile-OTP citizen account, not just a CAPTCHA.", ncols)
    ws.append([])
    ws.append(["Field", "Source", "Availability", "Cost", "Evidence", "Note"])
    header_row = ws.max_row
    _style_header(ws, header_row, ncols)
    for row in CTS_FINDINGS:
        ws.append(list(row))
        ws.cell(row=ws.max_row, column=3).fill = _STATUS_FILL.get(row[2], _STATUS_FILL["Unaudited"])
        ws.cell(row=ws.max_row, column=5).fill = _EVIDENCE_FILL.get(row[4], _EVIDENCE_FILL["unaudited"])
    _finish(ws, [34, 40, 14, 22, 16, 70], header_row, header_row + 1, ncols)


def build_cin_sheet(wb):
    ws = wb.create_sheet("C - CIN corporate")
    ncols = 7
    _title(ws, "Workflow C: CIN / corporate identity -- already national",
           "Keyed on CIN/LLPIN; every source is national. No state adapter needed.", ncols)
    ws.append([])
    ws.append(["Data item", "Source", "Availability", "Cost", "Evidence", "Charter facts field", "Note"])
    header_row = ws.max_row
    _style_header(ws, header_row, ncols)
    for row in CIN_FINDINGS:
        ws.append(list(row))
        ws.cell(row=ws.max_row, column=3).fill = _STATUS_FILL.get(row[2], _STATUS_FILL["Unaudited"])
        ws.cell(row=ws.max_row, column=5).fill = _EVIDENCE_FILL.get(row[4], _EVIDENCE_FILL["unaudited"])
    _finish(ws, [42, 34, 13, 14, 16, 30, 66], header_row, header_row + 1, ncols)


def build_unmapped_sheet(wb):
    ws = wb.create_sheet("D - Available but unused")
    ncols = 5
    _title(ws, "Available but NOT yet used in the Company Charter",
           "The point of this workbook: data we can already reach that no Charter field consumes.", ncols)
    ws.append([])
    ws.append(["Data", "Where it comes from", "Applies to", "What it would fix", "Value"])
    header_row = ws.max_row
    _style_header(ws, header_row, ncols)
    for row in UNMAPPED:
        ws.append(list(row))
        cell = ws.cell(row=ws.max_row, column=5)
        cell.fill = _STATUS_FILL["Yes"] if row[4] == "High" else _STATUS_FILL["Partial"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
    _finish(ws, [40, 30, 26, 64, 10], header_row, header_row + 1, ncols)


def build_readme(wb):
    ws = wb.active
    ws.title = "How to read this"
    ws.column_dimensions["A"].width = 118
    lines = [
        ("RERA data coverage", _TITLE_FONT),
        (f"Regenerated {date.today().isoformat()} by build_data_coverage.py", Font(italic=True, size=9, color="595959")),
        ("", None),
        ("What this is", Font(bold=True, size=12)),
        ("A record of what each authority and each workflow actually yields, so we can map later which "
         "of it the Company Charter should consume. Sheet D is the payload: data we can already reach "
         "that no Charter field uses yet.", None),
        ("", None),
        ("Do not hand-edit this file", Font(bold=True, size=12)),
        ("It is generated. FINDINGS in build_data_coverage.py is the source of truth -- edit there and "
         "rerun, or the next run silently discards your changes.", None),
        ("", None),
        ("Evidence levels -- read these before trusting a row", Font(bold=True, size=12)),
        ("confirmed-live   this pipeline actually fetched it", None),
        ("observed         seen on the portal, not yet fetched by code", None),
        ("unaudited        nobody has looked. ~20 of ~30 state portals are here, and so "
         "are individual rows on UP-RERA, TNRERA, HARERA and Delhi-RERA.", None),
        ("", None),
        ("Biggest findings so far", Font(bold=True, size=12)),
        ("1. GujRERA publishes AUDITED BALANCE SHEETS, P&L and income-tax returns per project. "
         "MahaRERA publishes nothing equivalent. Already downloaded by the adapter, not yet used.", None),
        ("2. The Maha Bhulekh land-record card already carries owner, area, tenure, encumbrance and "
         "mutation. None of it reaches the document -- the parser looks for the wrong HTML shape.", None),
        ("3. MCA charge filings are on the ZaubaCorp page the pipeline already fetches, and are the only "
         "independent check on a promoter's declared mortgage. Not parsed.", None),
        ("4. GujRERA does NOT link a promoter to their other projects -- confirmed live. So no promoter "
         "track record is possible there, and the capability is declared false.", None),
        ("5. K-RERA's PER-PROJECT complaint page is unreliable: it showed NO complaints for a project "
         "the state-wide register lists with 12. Counts must come from /projectComplaintReport. "
         "Reading the wrong page would have produced a false clean record.", None),
        ("6. K-RERA embeds its entire 9,888-project index client-side, giving a promoter-to-projects "
         "map for the whole state in one request -- the opposite of Gujarat.", None),
        ("7. MCA charge filings were on the ZaubaCorp page all along. A real promoter shows 4 OPEN "
         "charges totalling ~Rs 90.3 crore to HDFC Bank and Catalyst Trusteeship -- free, no extra "
         "request, and the only independent check on a declared mortgage.", None),
        ("8. Group membership CANNOT be derived from a brand name. Searching 'PRANAMI' returns a Delhi "
         "hydro-power firm, a Gujarat non-profit and a Maharashtra castings company. Propose by name, "
         "confirm by shared director / registered office / filed relationship -- and treat a shared "
         "address as the weakest of the three (28 of 65 'group companies' were address-only).", None),
        ("9. Two more states are live: Jharkhand (JHARERA) and West Bengal (WBRERA), taking coverage to "
         "six authorities at the time. JHARERA is the only authority anywhere in this pipeline that files a "
         "professional's PAN as a plain readable field rather than a scanned card, and it also publishes "
         "a declared past-projects table and a genuine litigation disclosure per project.", None),
        ("9b. FOUR more states are live as of 2026-08-24 -- Uttar Pradesh, Tamil Nadu, Haryana and "
         "Delhi -- taking coverage to TEN authorities. Each brings something no other one here "
         "does. HARERA states the promoter's CIN outright, which is the only hard RERA-to-MCA "
         "join in this pipeline; every other authority forces a name match. UP-RERA publishes "
         "the project bank account IN FULL and a khasra-level land grid. TNRERA publishes three "
         "separate order registers, all naming both parties. Delhi-RERA publishes almost nothing "
         "per project -- 130 projects for the whole NCT and no reachable detail record -- which "
         "is itself the finding: an absence from Delhi's register is close to worthless as "
         "evidence.", None),
        ("10. Karnataka's order/judgment coverage grew from one register to five (15,600+ rows across "
         "order search, authority orders, AO orders, interim orders and complaints-under-process), "
         "including a penalty table naming the violation, section and amount -- the single most "
         "consequential regulatory-history record any authority here publishes. JHARERA and WBRERA "
         "orders are now searchable by promoter too; WBRERA's join runs through its cause lists, "
         "since its own order register names no party directly.", None),
        ("11. The 32 'Unaudited' cells UP-RERA/TNRERA/HARERA/Delhi-RERA picked up on 2026-08-24 were "
         "actually audited live on 2026-08-26, one portal at a time, rather than left to sit. Each of "
         "the four states turned out to have a real Appellate Tribunal separate from the RERA authority "
         "itself -- HARERA's (HREAT) is plainly browsable and promoter-searchable, UP's (UP-REAT) is "
         "CAPTCHA-gated, Delhi's (REAT) is browsable but names no party, and Tamil Nadu's exists but "
         "publishes nothing searchable at all. Every one of the four also turned out to publish SOME "
         "form of defaulter/enforcement register the earlier audit had not found -- UP-RERA's "
         "DefaulterList (72 rows, reachable only via an ASP.NET postback, not a link), HARERA's "
         "cancelled/defaulter register (confirmed DISTINCT from its already-known lapsed-projects "
         "register, which is validity expiry rather than an authority action), TNRERA's penalty "
         "register (147 rows), and Delhi's execution register (7,493 rows) plus its suo-moto register "
         "(1,797 rows, the clearest 'projects under investigation' signal of the four). None of these "
         "seven new registers is wired into acquire() yet -- each earned an unwired, live-verified "
         "parser function instead, the same precedent WBRERA's fetch_defaulters() set, and each is its "
         "own row on Sheet D. What stayed genuinely negative also got a real reason instead of a shrug: "
         "none of the four publishes a P&L statement or an income-tax return anywhere, UP-RERA and "
         "Delhi-RERA publish no balance sheet either, and none but K-RERA states an NOC EXPIRY date "
         "(HARERA's 'Statutory Approvals' table gives the date obtained, not the date it lapses).",
         None),
        ("12. WBRERA's two 'Not built' rows are now built and confirmed live 2026-08-31. Land Area "
         "and a full location block (address/district/block/police station/pincode) turned out to be "
         "structured fields on every project page sampled -- adapter_westbengal.land_details() now "
         "reads them. A dag/mouza/J.L. survey reference is genuinely inconsistent, embedded as free "
         "text in the project address on roughly a third of a six-district sample and absent from the "
         "rest, hence 'Partial' rather than 'Yes'. The bank-account row went the other way: checked "
         "across the same six-project sample and confirmed the authority publishes no escrow/"
         "collection account field at all, so it moves from 'Not built' to 'Not published' -- a "
         "finding about WBRERA, not a gap in this adapter.", None),
        ("13. WBRERA's document library was already downloading a real audited balance sheet and up "
         "to three years of ITRs for every one of 12 projects sampled live 2026-08-31 -- the gap was "
         "that company_charter.py's own high-priority-document keyword list never recognised WBRERA's "
         "bare 'ITR' filenames ('ITR_3 Years.pdf', 'itr with blance sheet ay 23-4 & 24-25.pdf') as "
         "financial disclosure, so they were fetched but never text-extracted for the Charter pass. "
         "Fixed with an alpha-boundary regex rather than a plain substring (which would also fire on "
         "'arbitration'/'distribution'/'contribution') or a strict word boundary (which would miss the "
         "live 'ITR_3'/'ITR_23-24' filenames, since '_' counts as a word character). AUDITED BALANCE "
         "SHEET and Income-tax returns both move from No to Yes for WB.", None),
        ("14. Five of the sheet's 'Unaudited' cells were closed out live 2026-08-31. Professionals of "
         "record and Past-experience declarations are both genuinely Not published on HARERA -- its "
         "REP-I form has no section asking for either, confirmed by listing every numbered heading on "
         "the form across 7 real projects, not just the absence of a keyword. Past-experience is "
         "ALSO Not published on UP-RERA and TNRERA, checked on 4 and 3 real projects respectively -- "
         "a different finding from 'Promoter's other projects (track record)' above, which is about "
         "finding a promoter's OTHER filings by name match rather than a self-declared history on "
         "this one. Construction progress/QPR moved three ways at once: JHARERA to Yes (a genuine "
         "per-unit 'Completion Status' column already flowing through the adapter's own 'units' table "
         "but never surfaced), WBRERA to Not published (the page only carries a site-wide notice that "
         "QPR filing is mandatory, never a per-project figure), and HARERA to Partial (a real but "
         "project-type-dependent percentage-completion/status field). The JHARERA portal itself went "
         "down mid-audit -- 3 of 4 fetch attempts against the same URL timed out -- a reminder that "
         "an unreachable portal during a check is not the same finding as a confirmed absence.", None),
        ("15. Sheet B's last 'unaudited'/'Not built' row -- land records for the four non-Maharashtra "
         "states this pipeline touches -- is now four confirmed-live rows. Karnataka (Bhoomi) and "
         "Gujarat (AnyROR) are both free, District/Taluk(a)/village drill-downs behind a CAPTCHA text "
         "field, confirmed by inspecting the live DOM directly -- worth flagging because a first-pass "
         "automated read of both pages, working from page text alone, wrongly reported no CAPTCHA on "
         "either; the CAPTCHA element only exists once client-side JS renders it. AnyROR's own "
         "record-type dropdown carries 15 options, including a 'KNOW KHATA BY OWNER NAME' owner-name "
         "search. Uttar Pradesh splits across two separate systems: upbhulekh.gov.in's khatauni copy "
         "view is CAPTCHA-gated (plus a separate full-login flow for a fuller record, not audited "
         "here), while ekhasra.up.gov.in is a DIFFERENT system entirely for crop/khasra survey data, "
         "not ownership. Telangana is the one genuine surprise: dharani.telangana.gov.in, the system "
         "this codebase's own comments named, refuses every connection because Telangana retired "
         "Dharani for Bhu Bharati (bhubharati.telangana.gov.in) under a 2025 Act -- and Bhu Bharati's "
         "own public site routes every service through a mobile-OTP citizen login, no anonymous lookup "
         "path found anywhere, unlike the other three states' CAPTCHA-only public search. The two "
         "stale 'Dharani' comments this finding traces to (states/telangana.py, charter_document.py) "
         "were corrected in the same pass.", None),
        ("16. Sheet B's 'Lender, amount, satisfaction status' row said observed/'NOT currently parsed' "
         "months after it actually was -- Sheet C's own MCA-charge-filings row for the same underlying "
         "fact was confirmed live back in commit d0536e6, this one just never got updated to match. "
         "Reconciled 2026-09-01.", None),
        ("17. Credit-rating rationale documents -- WIRED IN 2026-09-01, not just found. CARE, India "
         "Ratings and Infomerics each embed a path to their own rationale document in a response this "
         "pipeline already fetches for the bare rating; none of the three were being extracted. "
         "Confirmed real by downloading/opening an actual document for each: Infomerics' carried a "
         "full 'Financials (Standalone)' table (Total Debt, Tangible Net Worth, EBITDA, PAT); CARE's "
         "carried real Total Income/Net Worth/Gearing figures; India Ratings renders its rationale as "
         "rich HTML rather than a PDF and went further still -- bookings, collections, cash flow from "
         "operations, forecast leverage ratios, a Strengths/Weaknesses key-rating-drivers section. Now "
         "wired into lookup_credit_rating's own return shape (rationale_url per agency), cited in the "
         "Charter's rating-comparison table, and registered as its own sources[] entry -- not left as "
         "an unused finding the way the WBRERA ITR gap was before it, too. ICRA stays a real, "
         "identified gap (a per-entity /Rationale/ShowRationaleReport?Id= mechanism exists, but "
         "resolving which Id belongs to which company needs more work); CRISIL is unchanged, since its "
         "own short inline excerpt predates this pass. This closes Sheet D's 'Credit-rating rationale "
         "documents' row -- removed from there since it is no longer unused.", None),
        ("18. IGR Maharashtra e-Search -- Sheet B's last two 'observed' rows, closed live 2026-09-01 "
         "with a REAL pulled result, not just a reachable form. freesearchigrservice.maharashtra.gov.in "
         "offers two free, CAPTCHA-gated search modes (Property Details: Year/District/Village-or-Area/"
         "Survey-CTS-Milkat-Gat-Plot No.; Document Number: Registration Type/District, all 37, genuinely "
         "statewide/SRO/Year/Doc No.). Document Number/SRO Mumbai 9 (Andheri)/2024/#100 returned a real "
         "registered Leave & License agreement -- seller and purchaser named, full property description, "
         "AND the actual consideration (Rs 44,100/month plus Rs 5,29,200 advance) -- all in the result "
         "row itself, no extra click. Genuine limitation found along the way, from the site's own FAQ: "
         "Power of Attorney and Will deeds are excluded from search entirely. Equitable mortgages share "
         "the identical mechanism and evidence -- a Notice of Intimation is just one more possible deed-"
         "name value, not a separately filterable category -- but no live mortgage example was actually "
         "pulled this pass, so that row confirms the channel, not a specific instance. Also found: the "
         "page's own 'CERSAI Search' button is an integrated postback on this same portal, not an "
         "external redirect. Added to Sheet D as a new high-value, unbuilt finding -- nothing in this "
         "pipeline calls this portal yet, and it is the only independent, party-named check on a "
         "promoter's registered dealings this pipeline has found for Maharashtra.", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=text)
        if font:
            ws.cell(row=i, column=1).font = font
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")


def main():
    wb = Workbook()
    build_readme(wb)
    build_rera_sheet(wb)
    build_cts_sheet(wb)
    build_cin_sheet(wb)
    build_unmapped_sheet(wb)
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[OK] {OUT_PATH}")
    print(f"     {len(RERA_FINDINGS)} RERA items x {len(RERA_STATES)} authorities, "
          f"{len(CTS_FINDINGS)} CTS, {len(CIN_FINDINGS)} CIN, {len(UNMAPPED)} unused")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
